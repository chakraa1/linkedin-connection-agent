"""
LinkedIn Connection Crew — 4-step pipeline for personalized outreach.

Step 1: Discover profiles    (Boolean search via boolean_search_agent + Playwright)
Step 2: Generate messages    (direct cached Anthropic calls, 8-rule quality gate)
Step 3: Human Excel review   (Shortlisted Yes/No)
Step 4: Send connections     (Playwright browser automation, explicit confirm)
"""
import json
import os
import re
import uuid
from datetime import date as _date
from pathlib import Path

import yaml
from crewai import Agent, Crew, Process, Task
from rich.console import Console
from rich.panel import Panel
from rich.prompt import Prompt

from linkedin_connection_agent.tools.browser_tool import LinkedInBrowser
# from linkedin_connection_agent.tools.pdf_tool import extract_pdf_text  # PDF disabled
from linkedin_connection_agent.utils.llm_factory import LLMFactory
from linkedin_connection_agent.utils.message_validator import OutreachQualityChecker
from linkedin_connection_agent.utils.scheduler import ConnectionScheduler

console = Console()

_BASE = Path(__file__).parent.parent.parent
AGENTS_CFG = yaml.safe_load((_BASE / "config/agents.yaml").read_text(encoding="utf-8"))
TASKS_CFG  = yaml.safe_load((_BASE / "config/tasks.yaml").read_text(encoding="utf-8"))
ICP_CFG    = yaml.safe_load((_BASE / "config/icp_config.yaml").read_text(encoding="utf-8"))

# ── Persona loading ───────────────────────────────────────────────────────────
# persona.yaml is the single user-editable input that replaces all hardcoded
# Anindya-specific content. Falls back to icp_config.yaml values if absent.
_PERSONA_FILE = _BASE / "config/persona.yaml"
_PERSONA: dict = yaml.safe_load(_PERSONA_FILE.read_text(encoding="utf-8")) if _PERSONA_FILE.exists() else {}

def _p(path: str, default: str = "") -> str:
    """Safely read a dot-path from _PERSONA, e.g. 'sender.name'."""
    val = _PERSONA
    for key in path.split("."):
        if not isinstance(val, dict):
            return default
        val = val.get(key, {})
    return str(val).strip() if val else default

_RUN_ID_FILE = Path("outputs/.current_run_id")

# ── System prompt builders ────────────────────────────────────────────────────
# All prompts are built once at module load from persona.yaml so the strings
# are stable for the entire session — Anthropic prompt caching still activates
# (1,024-token minimum on Sonnet). Variable content arrives in the user turn.

def _build_profile_analyzer_system() -> str:
    sender  = f"{_p('sender.name','the sender')}, {_p('sender.role')} at {_p('sender.company')}"
    bg      = _p("sender.background")
    goal    = _p("goal", "PEER_COLLABORATION")
    inds    = ", ".join(_PERSONA.get("target", {}).get("industries", ["Technology"]))
    kws     = ", ".join(_PERSONA.get("target", {}).get("keywords", []))
    return f"""\
You extract conversation hooks from LinkedIn profiles to help {sender}
achieve this goal: {goal}.

Sender background: {bg}

Target industries: {inds}
Target keywords: {kws}

Your output feeds directly into personalised connection messages — quality here
determines message quality downstream.

━━━ OUTPUT FORMAT (exact — no prose, no preamble, no deviation) ━━━

## RELEVANCE: [HIGH/MEDIUM/LOW]
## Rationale: [one specific line — what in the profile justifies this score]
## Top 3 Conversation Hooks
1. [HOOK_NAME]: [specific detail] — [why this anchors a genuine conversation]
2. [HOOK_NAME]: [specific detail] — [why this anchors a genuine conversation]
3. [HOOK_NAME]: [specific detail] — [why this anchors a genuine conversation]
## Recommended Hook: [1/2/3] — [one-line justification]

━━━ RELEVANCE SCORING ━━━

HIGH — strong outreach candidate:
  Clear seniority signal in title (VP, Director, Partner, CTO, CIO, Managing Director,
  Head of X, Founder, C-suite) AND substantive profile content (experience, about section).

MEDIUM — consider with lower priority:
  Moderate seniority. Some relevant content but limited detail.

LOW — skip for now:
  Junior titles, sparse profile, no clear connection to the sender's goal.

━━━ VALID HOOKS (must be specific and non-obvious) ━━━

A good hook is something specific that could anchor a genuine peer conversation.
It must be non-obvious — not just their title or company name.

Examples of valid hooks:
  - A specific challenge, transition, or decision visible in their profile
  - A named project, migration, launch, or organisational change with context
  - A career inflection point with specific numbers or outcomes
  - A topic they publicly write or speak about (if posts available)

Examples of what is NOT a hook:
  - Job title alone: "CTO at XYZ"
  - Company name alone: "works at Goldman Sachs"
  - Vague tenure: "8 years at HSBC — what did they build?"
  - Generic praise: "impressive career trajectory"

━━━ QUALITY STANDARD ━━━

The Recommended Hook should be the one most likely to trigger a thoughtful reply.
Ask yourself: "Would this person find this observation interesting enough to respond to?"

Profile name and data will be provided in the user message.\
"""


def _build_post_analyzer_system() -> str:
    sender  = f"{_p('sender.name','the sender')}, {_p('sender.role')} at {_p('sender.company')}"
    bg      = _p("sender.background")
    return f"""\
You identify the single best conversation entry point from a person's LinkedIn posts,
for use in a personalised connection request from {sender}.

Sender background: {bg}

━━━ OUTPUT FORMAT (exact — no deviation) ━━━

## Best Post Hook
Post reference: [brief description — which post, approx when]
Key observation/tension: [specific quote or close paraphrase from their post]
Conversation angle: [what the sender would say in response, given their background]
Confidence: HIGH or MEDIUM

Return exactly the string NO_POSTS_AVAILABLE if no posts are provided.

━━━ WHAT MAKES A HIGH-CONFIDENCE POST HOOK ━━━

A HIGH-confidence hook comes from a post where the person:
• States a non-obvious opinion or takes a position
• Surfaces a genuine tension or tradeoff
• Makes a specific prediction about their industry or domain
• Poses a question that reveals how they think about systemic problems

A MEDIUM-confidence hook: milestone or experience with some reflective insight.

━━━ WHAT DOES NOT MAKE A GOOD HOOK ━━━

Generic career announcements, reposts without commentary, celebration posts,
vague "leadership is about people" statements, job/hiring posts.

━━━ CALIBRATION ━━━

Choose the hook that gives the message writer the most specific starting point.
A direct quote beats a paraphrase. A tension beats an observation.
Specificity always beats generality.\
"""


def _build_message_writer_system() -> str:
    sender_name    = _p("sender.name", "the sender")
    sender_role    = _p("sender.role")
    sender_company = _p("sender.company")
    sender_bg      = _p("sender.background")
    goal           = _p("goal", "PEER_COLLABORATION")
    goal_desc      = _p("goal_description")
    tone_key       = _p("outreach_tone", "peer")
    msg_ctx        = _p("message_context")

    tone_map = {
        "peer":       "Collaborative peer energy. Curious about their work, not transactional.",
        "consultant": "Credible advisor who has seen this problem before. Specific, not pitching.",
        "job_seeker": "Curious about their work. Humble and specific. Never beg or hint at job search.",
        "recruiter":  "Respectful talent partner. Direct about value. Never waste their time.",
    }
    tone_instruction = tone_map.get(tone_key, tone_map["peer"])

    goal_framing_map = {
        "PEER_COLLABORATION":  "Show genuine curiosity about their work. No hidden agenda.",
        "CLIENT_ACQUISITION":  "Show curiosity about their challenge first. Do not pitch services.",
        "TALENT_ACQUISITION":  "Respect their current role. Be specific about why you are reaching out.",
        "JOB_HUNTING":         "Express genuine interest in their work. NEVER ask for a job or referral.",
        "INVESTOR_OUTREACH":   "Share a perspective relevant to their thesis. Peer-to-peer, not pitching.",
    }
    goal_framing = goal_framing_map.get(goal, "")

    return f"""\
Write a short, crisp LinkedIn outreach message on behalf of {sender_name},
{sender_role} at {sender_company}.

Sender background: {sender_bg}

Outreach goal: {goal}
{goal_desc}

{goal_framing}
{msg_ctx}

TONE: {tone_instruction}

━━━ LENGTH ━━━

80 to 150 words. Every word must earn its place. No padding.
Three short paragraphs maximum. Each paragraph: 1-3 sentences only.

━━━ POST RULE (critical — no hallucination) ━━━

If the profile context does NOT include a "Post Hook" section or says
"no public posts", the recipient has ZERO public LinkedIn posts.
Do NOT mention, reference, or imply any posts. Base the message entirely
on their role, company, experience, and the AI hooks provided.

━━━ PROFILE URL ━━━

The LinkedIn URL is provided. Reference something specific you found on
this exact profile. Not "came across your profile" — name what caught attention.

━━━ STRUCTURE (three tight paragraphs, no headers or bullets) ━━━

Para 1 — Hook (1-2 sentences): One sharp, specific observation about this person.
Must be specific enough that it could not be sent to anyone else.

Para 2 — Curiosity (1-2 sentences): One non-obvious question or tension
relevant to their background. Show you have thought about their actual problem.
This is the sentence that gets a reply.

Para 3 — Ask (1 sentence): A simple, direct ask ending with a genuine
question (?) or clear CTA. Nothing longer.

━━━ HARD RULES ━━━

A. Average sentence ≤ 12 words. Short sentences only.
B. 80-150 words. Hard fail outside this range.
C. Opening must be specific — cannot fit any other person unchanged.
D. Include one non-obvious insight or question relevant to their domain.
E. No "I built", "I led", "I launched", "we built", "we deployed".
F. No em-dashes (—), en-dashes (–), or double hyphens (--).
G. 2-3 distinct specific hooks. No vague phrases.
I. Last sentence must be a genuine question (?) or direct CTA.

━━━ NEVER ━━━

- Mention jobs, referrals, resumes, certifications, or years of experience
- "I hope this finds you well", "I wanted to reach out", "as someone who"
- "leverage", "synergize", "thought leadership", "game-changer", "deep dive"
- "amazing", "incredible", "brilliant", "impressive profile"
- Subject line, salutation, signature, or word count
- Any reference to posts if no Post Hook section exists

Return ONLY the message text. No labels, no quotes.\
"""


# Build all prompts once at module load — stable strings for prompt caching
_PROFILE_ANALYZER_SYSTEM = _build_profile_analyzer_system()
_POST_ANALYZER_SYSTEM     = _build_post_analyzer_system()
_MESSAGE_WRITER_SYSTEM    = _build_message_writer_system()

_REVISE_SYSTEM = """\
Revise a LinkedIn outreach message to fix the listed quality issues.
Keep the revised message between 80 and 150 words.
Short sentences only. Every word must earn its place.
Preserve whatever specific observation about the recipient already exists.
Do not replace a specific detail with a generic one.
Return ONLY the revised message. No explanation, no word count.\
"""

# Roles explicitly excluded — matched before inclusion check
_EXCLUDE_TITLE_KEYWORDS = [
    "junior",
    "associate vice president",
    "associate vp",
    "associate director",
    "associate partner",
    " associate ",      # standalone "Associate" role (e.g. "Associate, Goldman Sachs")
    "software engineer",
    "software developer",
    "project manager",
    "program manager",
    "product manager",
    "data analyst",
    "business analyst",
    # Note: "consultant" is intentionally NOT excluded here — McKinsey Senior Consultants,
    # BCG Associates, and Deloitte Consultants should pass so consulting firm queries work.
    # ICP scoring (STRONG/WEAK/MISMATCH) will filter irrelevant consultants downstream.
    "intern",
    "trainee",
    "analyst",          # plain "Analyst" without Senior/Lead prefix
]

# Leadership, Director-level and above roles that are included
_SENIOR_TITLE_KEYWORDS = [
    # C-Suite / Chief
    "ceo", "cto", "cio", "coo", "cfo", "cmo", "cpo", "chro", "cxo",
    "chief ",           # "Chief Digital Officer", "Chief of Staff", etc.
    # President / Vice President
    "president",
    "vice president",
    " vp ",             # "VP Engineering", "SVP", "EVP"
    "svp", "evp",
    # Director (all types)
    "managing director",
    "executive director",
    "senior director",
    "sr. director",
    "director",         # catches "Director of X", "Technical Director", etc.
    # Manager (only senior level)
    "senior manager",
    "sr. manager",
    "sr manager",
    "general manager",
    "country manager",
    "regional manager",
    # Head of / Global Head
    "head of ",
    "global head",
    "regional head",
    # Partner / Founder / Consulting senior titles
    "managing partner",
    "founding partner",
    "partner",
    "founder",
    "co-founder",
    "principal",          # McKinsey/BCG/Bain Principal = Director-equivalent
    "associate partner",  # Deloitte/EY/KPMG senior grade
    # Other leadership
    "md ",              # "MD | Goldman" or "MD Technology"
]


def _is_senior(headline: str) -> bool:
    h = " " + headline.lower() + " "
    hl = headline.lower()
    for kw in _EXCLUDE_TITLE_KEYWORDS:
        # Use word-boundary matching so "software engineer" doesn't exclude
        # "Director of Software Engineering" (a senior role).
        if _word_in(kw.strip(), hl):
            return False
    return any(kw in h for kw in _SENIOR_TITLE_KEYWORDS)


def _parse_relevance(hooks_text: str) -> str:
    """Extract HIGH / MEDIUM / LOW from analyzed hooks output."""
    if not hooks_text:
        return ""
    for line in hooks_text.split("\n"):
        if "RELEVANCE:" in line.upper():
            upper = line.upper()
            if "HIGH" in upper:
                return "HIGH"
            if "MEDIUM" in upper:
                return "MEDIUM"
            if "LOW" in upper:
                return "LOW"
    return ""


# Domain words that indicate a technical/engineering/fintech context in a headline
_TECH_DOMAIN_WORDS = [
    "engineering", "technology", "tech", "cloud", "platform", "infrastructure",
    "software", "data", "architecture", "digital", "devops", "fintech",
    "banking", "financial", "payments", "trading", "capital markets",
]


def _word_in(term: str, text: str) -> bool:
    """True if `term` appears as a whole word in `text` (case-insensitive, pre-lowercased)."""
    return bool(re.search(r'\b' + re.escape(term) + r'\b', text))


def _score_icp_fit(headline: str, about: str, experience_text: str, icp: dict) -> tuple[str, str]:
    """Score profile fit against ICP criteria using headline, About, and experience.

    Returns (label, reason) where label is 'STRONG', 'WEAK', or 'MISMATCH'.
    Priority: explicit role match > 2+ tech keywords > industry+keyword > domain word in headline.
    Uses word-boundary matching so short terms like 'cto' don't match inside 'director'.
    """
    # Combine all text; headline doubled to give it more weight
    full_text = (headline + " " + headline + " " + about + " " + experience_text).lower()
    headline_lower = headline.lower()

    keywords = [kw.lower() for kw in icp.get("keywords", [])]
    industries = [ind.lower() for ind in icp.get("industries", [])]
    roles = [r.lower() for r in icp.get("target_roles", [])]

    matched_roles = [r for r in roles if _word_in(r, full_text)]
    matched_kw = [kw for kw in keywords if _word_in(kw, full_text)]
    matched_ind = [ind for ind in industries if _word_in(ind, full_text)]
    headline_domain = [w for w in _TECH_DOMAIN_WORDS if _word_in(w, headline_lower)]

    # STRONG: explicit role title match
    if matched_roles:
        return "STRONG", f"Role match: {matched_roles[0]}"

    # STRONG: 2+ tech keywords present
    if len(matched_kw) >= 2:
        return "STRONG", f"Keywords: {', '.join(matched_kw[:3])}"

    # STRONG: industry + at least one tech keyword
    if matched_kw and matched_ind:
        return "STRONG", f"{matched_ind[0].title()} + {matched_kw[0]}"

    # WEAK: headline itself contains a domain word (e.g. "Director of Engineering")
    # — common for C-suite who don't fill in About with tech keywords
    if headline_domain:
        return "WEAK", f"Headline domain: {', '.join(headline_domain[:2])}"

    # WEAK: single keyword match in About
    if matched_kw:
        return "WEAK", f"Keyword (About only): {matched_kw[0]}"

    # WEAK: industry match only
    if matched_ind:
        return "WEAK", f"Industry only: {matched_ind[0].title()}"

    return "MISMATCH", "No ICP signals in headline or About/Experience"


def _load_or_create_run_id(new_run: bool = False) -> str:
    """Persist run ID in outputs/.current_run_id so all pipeline steps share one Excel."""
    if new_run or not _RUN_ID_FILE.exists():
        run_id = uuid.uuid4().hex[:8]
        _RUN_ID_FILE.parent.mkdir(parents=True, exist_ok=True)
        _RUN_ID_FILE.write_text(run_id)
        return run_id
    content = _RUN_ID_FILE.read_text().strip()
    return content if content else _load_or_create_run_id(new_run=True)


class LinkedInConnectionCrew:
    def __init__(self):
        self._llm = LLMFactory()
        self._scheduler = ConnectionScheduler()
        self._quality_checker = OutreachQualityChecker()
        self._run_id = _load_or_create_run_id(new_run=False)
        self._anthropic = None  # lazy-init Anthropic client

    @property
    def _client(self):
        if self._anthropic is None:
            from anthropic import Anthropic
            self._anthropic = Anthropic()
        return self._anthropic

    # ------------------------------------------------------------------ #
    # Direct cached Anthropic helpers  (bypass CrewAI per-profile overhead)
    # ------------------------------------------------------------------ #

    def _analyze_profile_direct(self, profile_data_str: str, profile_name: str) -> str:
        """Profile relevance + hooks via cached system prompt. ~250 tokens/call vs ~850 before."""
        resp = self._client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=600,
            system=[{"type": "text", "text": _PROFILE_ANALYZER_SYSTEM,
                     "cache_control": {"type": "ephemeral"}}],
            messages=[{"role": "user", "content": (
                f"Profile name: {profile_name}\n\nProfile data:\n{profile_data_str[:3000]}"
            )}],
        )
        return resp.content[0].text.strip()

    def _analyze_posts_direct(self, posts_text: str) -> str:
        """Post hook extraction via cached system prompt.
        Uses Sonnet (1,024-token threshold) since the system prompt is ~1,100 tokens.
        Haiku requires 2,048 tokens minimum — impractical for this prompt size.
        """
        resp = self._client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=300,
            system=[{"type": "text", "text": _POST_ANALYZER_SYSTEM,
                     "cache_control": {"type": "ephemeral"}}],
            messages=[{"role": "user", "content": f"Posts:\n{posts_text[:2000]}"}],
        )
        return resp.content[0].text.strip()

    def _write_message_direct(self, profile_context: str) -> str:
        """Write an 80-150 word crisp outreach message using cached system prompt."""
        resp = self._client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=250,  # 150 words ≈ 200 tokens; small headroom
            system=[{"type": "text", "text": _MESSAGE_WRITER_SYSTEM,
                     "cache_control": {"type": "ephemeral"}}],
            messages=[{"role": "user", "content": (
                f"Profile data:\n{profile_context[:2200]}"
            )}],
        )
        return resp.content[0].text.strip()

    # ------------------------------------------------------------------ #
    # Phase 1: Boolean search string loading / generation
    # ------------------------------------------------------------------ #

    _SEARCH_CONFIG = Path("config/search_strings.yaml")

    def _load_search_strings(
        self, icp_key: str, locations_override: list[str] | None
    ) -> list[dict]:
        """Priority: search_strings.yaml → generate from icp_config.yaml → auto-save YAML.

        1. If config/search_strings.yaml exists and has entries → use those directly.
           (Skips tasks.yaml and the LLM entirely — zero token cost.)
        2. If the file is absent → generate via LLM using tasks.yaml + icp_config.yaml data.
        3. Auto-save the LLM output to config/search_strings.yaml so subsequent runs
           use the cached file and never regenerate unnecessarily.
        """
        if self._SEARCH_CONFIG.exists():
            cfg = yaml.safe_load(self._SEARCH_CONFIG.read_text(encoding="utf-8"))
            strings = cfg.get("search_strings", [])
            if strings:
                console.print(
                    f"[dim]Loaded {len(strings)} curated queries from "
                    f"{self._SEARCH_CONFIG}[/dim]"
                )
                by_segment: dict[str, int] = {}
                for s in strings:
                    seg = s.get("segment", "unknown")
                    by_segment[seg] = by_segment.get(seg, 0) + 1
                for seg, count in by_segment.items():
                    console.print(f"  [dim]{seg}: {count} quer{'y' if count == 1 else 'ies'}[/dim]")
                return strings

        console.print(
            "[dim]config/search_strings.yaml not found — "
            "generating from icp_config.yaml via tasks.yaml...[/dim]"
        )
        strings = self.generate_search_strings(icp_key, locations_override=locations_override)
        self._save_search_config(strings, source=f"icp_config.yaml (icp: {icp_key})")
        return strings

    def _save_search_config(self, strings: list[dict], source: str = "icp_config.yaml") -> None:
        """Persist search strings to config/search_strings.yaml."""
        doc = {
            "_source": source,
            "_note": (
                "Auto-saved from icp_config.yaml. "
                "Edit queries here to fix them for future runs, or delete this file to regenerate."
            ),
            "search_strings": strings,
        }
        self._SEARCH_CONFIG.parent.mkdir(parents=True, exist_ok=True)
        self._SEARCH_CONFIG.write_text(
            yaml.dump(doc, allow_unicode=True, sort_keys=False, width=120, default_flow_style=False),
            encoding="utf-8",
        )
        console.print(
            f"[dim]Auto-saved {len(strings)} queries → {self._SEARCH_CONFIG} "
            f"(reused on next run)[/dim]"
        )

    def generate_search_strings(
        self, icp_key: str = "icp1", locations_override: list[str] | None = None
    ) -> list[dict]:
        # Prefer persona.yaml over icp_config.yaml when persona exists
        icp = ICP_CFG.get(icp_key, {})
        persona_target = _PERSONA.get("target", {})
        persona_segments = _PERSONA.get("segments", [])

        locations = (
            locations_override
            or persona_target.get("locations")
            or icp.get("locations", ["India"])
        )
        target_roles = persona_target.get("roles") or icp.get("target_roles", [])
        industries   = persona_target.get("industries") or icp.get("industries", [])
        keywords     = persona_target.get("keywords") or icp.get("keywords", [])
        target_desc  = persona_target.get("description") or icp.get("target_profile_description", "")
        segments_str = json.dumps(
            [{"name": s["name"], "description": s["description"]} for s in persona_segments]
            if persona_segments else [{"name": "General", "description": "All ICP targets"}]
        )

        agent = Agent(
            role=AGENTS_CFG["boolean_search_agent"]["role"].format(
                sender_name=_p("sender.name", "the sender"),
                sender_role=_p("sender.role"),
                sender_company=_p("sender.company"),
                goal=_p("goal", "PEER_COLLABORATION"),
                goal_description=_p("goal_description"),
                segments=segments_str,
            ),
            goal=AGENTS_CFG["boolean_search_agent"]["goal"].format(
                sender_name=_p("sender.name", "the sender"),
                sender_role=_p("sender.role"),
                sender_company=_p("sender.company"),
                goal=_p("goal", "PEER_COLLABORATION"),
                goal_description=_p("goal_description"),
                segments=segments_str,
            ),
            backstory=AGENTS_CFG["boolean_search_agent"]["backstory"],
            llm=self._llm.get("boolean_search_agent"),
            verbose=False,
        )
        task = Task(
            description=TASKS_CFG["generate_boolean_search_task"]["description"].format(
                sender_name=_p("sender.name", "the sender"),
                sender_role=_p("sender.role"),
                sender_company=_p("sender.company"),
                goal=_p("goal", "PEER_COLLABORATION"),
                goal_description=_p("goal_description"),
                target_description=target_desc,
                target_roles=json.dumps(target_roles),
                industries=json.dumps(industries),
                locations=json.dumps(locations),
                keywords=json.dumps(keywords),
                segments=segments_str,
            ),
            expected_output=TASKS_CFG["generate_boolean_search_task"]["expected_output"],
            agent=agent,
        )
        crew = Crew(agents=[agent], tasks=[task], process=Process.sequential, verbose=False)
        output = str(crew.kickoff())
        try:
            start, end = output.find("["), output.rfind("]") + 1
            return json.loads(output[start:end])
        except Exception:
            return [{"query": output.strip(), "rationale": "raw output"}]

    # ------------------------------------------------------------------ #
    # Live query testing — verifies result counts before committing
    # ------------------------------------------------------------------ #

    @staticmethod
    def _count_query_results(browser: "LinkedInBrowser", query: str) -> int:
        """Open the search URL and count result cards. Returns 0 on any error."""
        import time as _time
        try:
            url = LinkedInBrowser._build_search_url(query)
            browser._page.goto(url, wait_until="load")
            _time.sleep(3)
            browser._page.evaluate("window.scrollTo(0, 400)")
            _time.sleep(1)
            if "login" in browser._page.url:
                return -1  # session expired
            cards = (
                browser._page.query_selector_all(".reusable-search__result-container")
                or browser._page.query_selector_all("li.artdeco-list__item")
            )
            if cards:
                return len(cards)
            links = browser._page.query_selector_all("a[href*='/in/']")
            seen: set[str] = set()
            for lnk in links:
                href = (lnk.get_attribute("href") or "").split("?")[0].rstrip("/")
                if "/in/" in href:
                    seen.add(href)
            return len(seen)
        except Exception:
            return 0

    def _filter_working_queries(
        self, browser: "LinkedInBrowser", strings: list[dict], min_results: int = 5
    ) -> list[dict]:
        """Test each query live and return only those with >= min_results cards."""
        import time as _time
        console.print(f"\n[dim]Auto-testing {len(strings)} queries (keeping ≥{min_results} cards)...[/dim]")
        working: list[dict] = []
        for s in strings:
            q = s.get("query", "")
            count = self._count_query_results(browser, q)
            if count == -1:
                console.print("[yellow]  Session expired during query test — stopping test.[/yellow]")
                break
            tag = "[green]OK[/green]" if count >= min_results else "[red]NO_RESULTS[/red]"
            console.print(f"  {tag} [{count:>2} cards] {q[:70]}...")
            if count >= min_results:
                working.append(s)
            _time.sleep(2)

        if not working:
            console.print("[yellow]  No queries passed the test. Using all queries unfiltered.[/yellow]")
            return strings

        dropped = len(strings) - len(working)
        if dropped:
            console.print(
                f"[dim]  {dropped} queries returned < {min_results} cards and were dropped.[/dim]"
            )
        return working

    # ------------------------------------------------------------------ #
    # Phase 2: Profile discovery
    # ------------------------------------------------------------------ #

    def discover_profiles(
        self, icp_key: str = "icp1", max_per_query: int = 10, location: str | None = None
    ) -> int:
        # Each discover run starts a fresh Excel with a new run ID
        self._run_id = _load_or_create_run_id(new_run=True)

        locations_override = [location] if location else None
        # Prefer persona.yaml locations over icp_config.yaml
        persona_locations = _PERSONA.get("target", {}).get("locations")
        active_locations = (
            locations_override
            or persona_locations
            or ICP_CFG.get(icp_key, {}).get("locations", ["India"])
        )
        console.print(f"[dim]Region filter: {', '.join(active_locations)}[/dim]")

        # Merge persona target fields into icp_config for ICP scoring
        persona_target = _PERSONA.get("target", {})
        base_icp = ICP_CFG.get(icp_key, {})
        icp_config = {
            "target_roles": persona_target.get("roles") or base_icp.get("target_roles", []),
            "industries":   persona_target.get("industries") or base_icp.get("industries", []),
            "keywords":     persona_target.get("keywords") or base_icp.get("keywords", []),
        }

        search_strings = self._load_search_strings(icp_key, locations_override)
        console.print(f"\n[bold cyan]{len(search_strings)} search strings ready.[/bold cyan]")
        for i, s in enumerate(search_strings, 1):
            seg = s.get("segment", "")
            label = f"[dim]{seg}[/dim] " if seg else ""
            console.print(f"  {i}. {label}{s['query'][:80]}...")
        skipped_junior = 0
        skipped_existing = 0

        # Fetch 2× max_per_query from LinkedIn per query so we have enough candidates
        # to select the top STRONG ICP matches after scoring — not just the first N returned.
        search_fetch_limit = min(max_per_query * 2, 15)

        pending_scrape: list[dict] = []   # new senior profiles not yet in DB
        seen_urls: set[str] = set()       # cross-query dedup within this run

        with LinkedInBrowser(headless=False) as browser:
            if not browser.is_logged_in():
                console.print("[yellow]Session expired — logging in with .env credentials...[/yellow]")
                if not browser.login(os.environ["LINKEDIN_EMAIL"], os.environ["LINKEDIN_PASSWORD"]):
                    console.print("[red]Login failed. Check LINKEDIN_EMAIL / LINKEDIN_PASSWORD in .env[/red]")
                    return 0

            # ── Auto-test: drop queries that return zero results before scraping ──
            # This prevents wasting scrape time on broken boolean queries.
            # Only runs when search_strings.yaml was just generated (not pre-tested).
            if not self._SEARCH_CONFIG.exists():
                search_strings = self._filter_working_queries(browser, search_strings)
                if search_strings:
                    self._save_search_config(search_strings, source=f"persona.yaml → tested live")

            # ── Phase 1: collect search results (fetch 2× for top-match filtering) ──
            session_refreshed = False
            for search in search_strings:
                console.print(f"\n[dim]Searching: {search['query'][:60]}...[/dim]")
                try:
                    profiles = browser.search_people(search["query"], max_results=search_fetch_limit)
                except RuntimeError as exc:
                    if "session expired" in str(exc) and not session_refreshed:
                        console.print("[yellow]Session expired mid-run — re-logging in...[/yellow]")
                        if not browser.login(os.environ["LINKEDIN_EMAIL"], os.environ["LINKEDIN_PASSWORD"]):
                            console.print("[red]Re-login failed.[/red]")
                            break
                        session_refreshed = True
                        profiles = browser.search_people(search["query"], max_results=search_fetch_limit)
                    else:
                        console.print(f"[red]Search failed: {exc}[/red]")
                        continue

                for p in profiles:
                    headline = p.get("headline", "") or ""

                    # Only hard-filter when we have a headline AND it's clearly junior.
                    # Empty headlines (LinkedIn fallback scrape) pass through to Phase 2
                    # where we scrape the real headline from the profile page.
                    if headline and not _is_senior(headline):
                        skipped_junior += 1
                        continue

                    url = p["url"]
                    if url in seen_urls:
                        continue
                    seen_urls.add(url)

                    if self._scheduler.get_by_url(url):
                        skipped_existing += 1
                        continue

                    pending_scrape.append(p)

            # ── Phase 2: scrape, ICP score, keep top STRONG matches ─────────────
            # Collect all scored profiles then prefer STRONG over WEAK up to the
            # target count (max_per_query × number of queries).
            strong_profiles: list[tuple] = []
            weak_profiles: list[tuple] = []
            skipped_mismatch = 0

            if pending_scrape:
                console.print(
                    f"\n[dim]Scraping {len(pending_scrape)} candidates — "
                    f"keeping top {max_per_query} STRONG ICP matches per query...[/dim]"
                )
                for p in pending_scrape:
                    console.print(f"  Scraping: [cyan]{p['name']}[/cyan]")
                    try:
                        scraped = browser.scrape_profile(p["url"])
                    except Exception as exc:
                        console.print(f"    [yellow]Scrape failed: {exc} — skipping.[/yellow]")
                        continue

                    full_headline = scraped.get("headline") or p.get("headline", "")

                    # Re-check seniority on the scraped headline for profiles that had
                    # empty search-result headlines (LinkedIn fallback path returns no headline)
                    if not _is_senior(full_headline):
                        skipped_junior += 1
                        console.print(f"    [dim]Skipping — not senior after scraping ({full_headline[:60] or 'no headline'})[/dim]")
                        continue

                    about_text = scraped.get("about", "") or ""
                    exp_list = scraped.get("experience", []) or []
                    exp_text = " ".join(str(e) for e in exp_list)

                    icp_fit, icp_reason = _score_icp_fit(
                        full_headline, about_text, exp_text, icp_config
                    )
                    fit_color = {"STRONG": "green", "WEAK": "yellow", "MISMATCH": "red"}.get(icp_fit, "dim")
                    console.print(f"    ICP Fit: [{fit_color}]{icp_fit}[/{fit_color}] — {icp_reason}")

                    if icp_fit == "MISMATCH":
                        skipped_mismatch += 1
                        continue

                    entry = (p, scraped, full_headline, about_text, exp_list, icp_fit, icp_reason)
                    if icp_fit == "STRONG":
                        strong_profiles.append(entry)
                    else:
                        weak_profiles.append(entry)

            # Select top matches: STRONG first, fill with WEAK only if under target
            target_count = max_per_query * len(search_strings)
            to_save = strong_profiles[:target_count]
            remaining_slots = max(0, target_count - len(to_save))
            if remaining_slots:
                to_save.extend(weak_profiles[:remaining_slots])

            console.print(
                f"\n[dim]Saving {len(to_save)} top matches "
                f"({len(strong_profiles)} STRONG available, {len(weak_profiles)} WEAK available, "
                f"{skipped_mismatch} MISMATCH discarded)[/dim]"
            )

            for p, scraped, full_headline, about_text, exp_list, icp_fit, icp_reason in to_save:
                profile_data = json.dumps({
                    "name": scraped.get("name") or p["name"],
                    "headline": full_headline,
                    "url": p["url"],
                    "about": about_text,
                    "experience": exp_list,
                    "icp_fit": icp_fit,
                    "icp_reason": icp_reason,
                }, indent=2)

                self._scheduler.save_discovered(
                    profile_url=p["url"],
                    profile_name=p["name"],
                    profile_headline=p.get("headline", ""),
                    icp_key=icp_key,
                    profile_data=profile_data,
                )

        new_count = len(to_save)
        console.print(
            f"\n[bold green]Discovered {new_count} ICP-matching profiles.[/bold green]"
            f"  [dim](skipped {skipped_junior} junior titles, "
            f"{skipped_existing} already in pipeline, "
            f"{skipped_mismatch} ICP mismatches)[/dim]"
        )

        # Export to Excel immediately — About + Experience now populated
        path = self.export_to_excel()
        console.print(f"[bold cyan]Discovery snapshot → {path}[/bold cyan]")
        return new_count

    # ------------------------------------------------------------------ #
    # Phase 3: Profile analysis (PDF-first, no posts)
    # ------------------------------------------------------------------ #

    def analyze_profiles(self, limit: int = 10) -> int:
        discovered = self._scheduler.list_by_status("discovered")[:limit]
        if not discovered:
            console.print("[yellow]No discovered profiles to analyze.[/yellow]")
            return 0

        # Phase A — browser: scrape posts only (About+Experience already in DB from discovery)
        scraped_data: list[tuple] = []   # (record, profile_data_str, posts)
        console.print("[dim]Phase 1/2: Scraping recent posts...[/dim]")
        with LinkedInBrowser(headless=False) as browser:
            if not browser.is_logged_in():
                console.print("[yellow]Session expired — logging in...[/yellow]")
                if not browser.login(os.environ["LINKEDIN_EMAIL"], os.environ["LINKEDIN_PASSWORD"]):
                    console.print("[red]Login failed.[/red]")
                    return 0

            for record in discovered:
                console.print(f"  Posts: [cyan]{record.profile_name}[/cyan]", end="")

                # Reuse About+Experience from discovery; only re-scrape if missing
                profile_data_str = record.profile_data or ""
                if not profile_data_str:
                    console.print(" [dim](re-scraping profile — not cached)[/dim]")
                    try:
                        scraped = browser.scrape_profile(record.profile_url)
                    except Exception:
                        scraped = {}
                    profile_data_str = json.dumps({
                        "name": scraped.get("name") or record.profile_name,
                        "headline": scraped.get("headline") or record.profile_headline,
                        "url": record.profile_url,
                        "about": scraped.get("about", ""),
                        "experience": scraped.get("experience", []),
                    }, indent=2)
                else:
                    console.print(" [dim](profile cached)[/dim]")

                # Stage 2.2 feed: scrape recent posts
                posts = browser.scrape_recent_posts(record.profile_url)
                post_label = f"[green]{len(posts)} posts[/green]" if posts else "[dim]no posts[/dim]"
                console.print(f"    → {post_label}")
                scraped_data.append((record, profile_data_str, posts))

        # Phase B — direct cached Anthropic calls (no CrewAI overhead per profile)
        # System prompts are cached after the first call → ~10% cost for all subsequent profiles.
        console.print("\n[dim]Phase 2/2: Running AI analysis (cached prompts)...[/dim]")

        processed = 0
        for record, profile_data_str, posts in scraped_data:
            console.print(f"  Analyzing: [cyan]{record.profile_name}[/cyan]")

            # Stage 2.1 — Relevance scoring + profile hooks
            analyzed_hooks = self._analyze_profile_direct(profile_data_str, record.profile_name)

            # Stage 2.2 — Post hook (only when active posts exist)
            if posts:
                post_text = "\n\n---\n\n".join(posts)
                post_hook_out = self._analyze_posts_direct(post_text)
                if post_hook_out and "NO_POSTS_AVAILABLE" not in post_hook_out:
                    analyzed_hooks = analyzed_hooks + "\n\n## Post Hook\n" + post_hook_out
                    console.print("    [green]Post hook appended.[/green]")
            else:
                console.print("    [dim]Stage 2.2 skipped — no posts.[/dim]")

            self._scheduler.save_analyzed(
                profile_id=record.id,
                profile_data=profile_data_str,
                recent_posts=analyzed_hooks,
                pdf_path="",
            )
            relevance = _parse_relevance(analyzed_hooks)
            rel_color = {"HIGH": "green", "MEDIUM": "yellow", "LOW": "red"}.get(relevance, "dim")
            console.print(f"    → [{rel_color}]RELEVANCE: {relevance or 'unknown'}[/{rel_color}]")
            processed += 1

        console.print(f"\n[bold green]Analyzed {processed} profiles.[/bold green]")

        # Re-export Excel: hooks, relevance, and auto-shortlist updated
        path = self.export_to_excel()
        console.print(f"[bold cyan]Analysis snapshot → {path}[/bold cyan]")
        return processed

    # ------------------------------------------------------------------ #
    # Phase 4 & 5: Message generation + validation
    # ------------------------------------------------------------------ #

    def generate_messages(self, limit: int = 10) -> int:
        # Process discovered OR analyzed profiles — analyze step is now optional
        candidates = (
            self._scheduler.list_by_status("analyzed")
            + self._scheduler.list_by_status("discovered")
        )[:limit]
        if not candidates:
            console.print("[yellow]No profiles ready for message generation.[/yellow]")
            return 0

        console.print(f"[dim]Writing messages for {len(candidates)} profiles...[/dim]")
        generated = 0

        for record in candidates:
            console.print(f"\n  Writing for: [cyan]{record.profile_name}[/cyan]")

            # Build profile context — include URL, about, experience, and AI analysis
            try:
                pd = json.loads(record.profile_data or "{}")
                about = pd.get("about", "") or ""
                exp_list = pd.get("experience", []) or []
                experience = "\n".join(str(e) for e in exp_list[:4])
            except Exception:
                about = ""
                experience = ""

            profile_context = (
                f"LinkedIn URL: {record.profile_url}\n"
                f"Name: {record.profile_name}\n"
                f"Headline: {record.profile_headline or ''}\n"
                f"About: {about[:600]}\n"
                f"Experience:\n{experience[:400]}\n"
            )

            # Append AI analysis (hooks + post hook if available)
            # This tells the writer exactly what is known and whether posts exist
            if record.recent_posts:
                profile_context += f"\nAI Analysis (profile hooks and post data):\n{record.recent_posts[:800]}"
            else:
                profile_context += "\n[No post analysis — recipient has no public posts. Do not reference any posts.]"

            message = self._write_message_direct(profile_context)

            # Full 8-rule quality gate (A, B, C, D, E, F, G, I)
            quality = self._quality_checker.check(message)

            if not quality.passed:
                rule_labels = [i.split(":")[0] for i in quality.issues]
                console.print(
                    f"    [yellow]Quality issues: {', '.join(rule_labels)}[/yellow]"
                )
                for issue in quality.issues:
                    console.print(f"      [dim]• {issue}[/dim]")

                # Auto-revise up to 2 times
                for attempt in range(2):
                    message = self._revise_message(message, quality.issues)
                    quality = self._quality_checker.check(message)
                    if quality.passed:
                        console.print(
                            f"    [green]Revision {attempt + 1} passed all rules.[/green]"
                        )
                        break
                    console.print(
                        f"    [yellow]Revision {attempt + 1}: still "
                        f"{len(quality.issues)} issue(s)[/yellow]"
                    )

                if not quality.passed:
                    remaining = [i.split(":")[0] for i in quality.issues]
                    console.print(
                        f"    [dim]Saving with {len(quality.issues)} unresolved "
                        f"({', '.join(remaining)}) — review manually[/dim]"
                    )

            console.print(f"    [green]{quality.word_count} words[/green]")
            self._scheduler.save_message(record.id, message)
            generated += 1

        console.print(f"\n[bold green]Messages written for {generated} profiles.[/bold green]")
        path = self.export_to_excel()
        console.print(f"[bold cyan]Saved → {path}[/bold cyan]")
        return generated

    # ------------------------------------------------------------------ #
    # Excel export / import
    # ------------------------------------------------------------------ #

    def export_to_excel(self, out_path: str | None = None) -> str:
        """4-column Excel: Name | LinkedIn URL | Message | Shortlisted."""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        if out_path is None:
            today = _date.today().strftime("%Y-%m-%d")
            out_path = f"outputs/{today}/profiles_review_{self._run_id}.xlsx"

        HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        MSG_FILL    = PatternFill(start_color="EFF7FF", end_color="EFF7FF", fill_type="solid")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Outreach"

        headers    = ["Name",  "LinkedIn URL", "Generated Message",  "Shortlisted (Yes / No)"]
        col_widths = [30,       52,              95,                   20]

        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 28

        for row_idx, profile in enumerate(self._scheduler.list_all_records(), 2):
            url = profile.profile_url or ""

            if profile.status in ("approved", "sent"):
                shortlisted = "Yes"
            elif profile.status == "rejected":
                shortlisted = "No"
            else:
                shortlisted = ""

            row = [profile.profile_name or "", url, profile.message or "", shortlisted]

            for col, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                if col == 2 and url:
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")
                    cell.alignment = Alignment(vertical="top")
                elif col == 3:
                    cell.fill = MSG_FILL
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            ws.row_dimensions[row_idx].height = 130

        ws.freeze_panes = "A2"
        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out))
        return str(out)

    def import_excel_review(self, path: str | None = None) -> tuple[int, int]:
        """
        Read profiles_review Excel (auto-detects most recent if path not specified).
        Shortlisted = 'Yes'  → approve (and update message if edited).
        Shortlisted = 'No'   → reject.
        Blank                → leave as-is.
        Returns (approved_count, rejected_count).
        """
        import openpyxl

        xlsx_path: Path
        if path:
            xlsx_path = Path(path)
        else:
            candidates = sorted(
                Path("outputs").rglob("profiles_review_*.xlsx"),
                key=lambda f: f.stat().st_mtime,
                reverse=True,
            )
            if candidates:
                xlsx_path = candidates[0]
            else:
                xlsx_path = Path("outputs/profiles_review.xlsx")

        if not xlsx_path.exists():
            raise FileNotFoundError(
                f"No review Excel found at {xlsx_path}. Run: python main.py export"
            )

        console.print(f"[dim]Importing from: {xlsx_path}[/dim]")
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb.active

        approved = rejected = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            cols = list(row)
            n = len(cols)

            # Current 4-col format: Name(0) URL(1) Message(2) Shortlisted(3)
            if n >= 4 and (n < 10 or str(cols[1] or "").startswith("http")):
                url         = cols[1]
                message     = cols[2]
                shortlisted = cols[3]
                notes       = ""
            elif n >= 15:
                url, message, shortlisted, notes = cols[2], cols[8], cols[13], cols[14]
            elif n >= 13:
                col8_val = str(cols[8] or "").strip().upper()
                if col8_val in ("PASS", "FAIL", ""):
                    url, message, shortlisted, notes = cols[2], cols[6], cols[11], cols[12]
                else:
                    url, message, shortlisted, notes = cols[2], cols[8], cols[11], cols[12]
            elif n >= 10:
                url, message, shortlisted, notes = cols[2], cols[5], cols[8], cols[9]
            else:
                continue

            if not url:
                continue
            profile = self._scheduler.get_by_url(str(url).strip())
            if not profile:
                continue
            if profile.status == "sent":
                continue

            val = str(shortlisted or "").strip().lower()
            if val == "yes":
                if message and str(message).strip() != (profile.message or "").strip():
                    self._scheduler.save_message(profile.id, str(message).strip())
                self._scheduler.approve_message(profile.id)
                approved += 1
            elif val == "no":
                self._scheduler.reject_message(profile.id, feedback=str(notes or ""))
                rejected += 1

        # Refresh Excel so Status and Shortlisted columns reflect import decisions
        path = self.export_to_excel()
        console.print(f"[bold cyan]Excel updated → {path}[/bold cyan]")
        return approved, rejected


    def _revise_message(self, message: str, issues: list[str]) -> str:
        issues_str = "\n".join(f"  • {i}" for i in issues)
        resp = self._client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=500,
            system=_REVISE_SYSTEM,
            messages=[{"role": "user", "content": (
                f"Original message:\n{message}\n\n"
                f"Fix every issue listed below. Keep the 200-300 word target.\n\n"
                f"Issues:\n{issues_str}"
            )}],
        )
        return resp.content[0].text.strip()

    # ------------------------------------------------------------------ #
    # Phase 6: Human review
    # ------------------------------------------------------------------ #

    def review_messages(self) -> int:
        drafts = self._scheduler.list_by_status("message_drafted")
        if not drafts:
            console.print("[yellow]No messages pending review.[/yellow]")
            return 0

        approved = 0
        for record in drafts:
            console.print("\n" + "=" * 60)
            console.print(Panel(
                f"[bold]{record.profile_name}[/bold]\n"
                f"[dim]{record.profile_headline}[/dim]\n"
                f"[cyan]{record.profile_url}[/cyan]",
                title="[bold cyan]Profile[/bold cyan]",
            ))

            if record.recent_posts:
                # Show first 400 chars of the hooks/analysis for context
                hooks_preview = record.recent_posts[:400].strip()
                if hooks_preview:
                    console.print(Panel(
                        hooks_preview + ("..." if len(record.recent_posts) > 400 else ""),
                        title="[dim]Profile Hooks (excerpt)[/dim]",
                    ))

            console.print(Panel(
                f"[bold white]{record.message}[/bold white]\n\n"
                f"[dim]{len(record.message or '')} / 300 characters[/dim]",
                title="[bold green]Proposed Message[/bold green]",
            ))

            action = Prompt.ask(
                "Action",
                choices=["approve", "edit", "skip", "reject"],
                default="approve",
            )

            if action == "approve":
                self._scheduler.approve_message(record.id)
                approved += 1
                console.print("[green]Approved.[/green]")
            elif action == "edit":
                new_msg = Prompt.ask("Enter revised message (300 chars max)")
                if len(new_msg) > 300:
                    console.print(f"[red]Too long ({len(new_msg)} chars). Skipping.[/red]")
                    continue
                self._scheduler.save_message(record.id, new_msg)
                self._scheduler.approve_message(record.id, feedback="human-edited")
                approved += 1
                console.print("[green]Edited and approved.[/green]")
            elif action == "reject":
                feedback = Prompt.ask("Rejection reason (optional)", default="")
                self._scheduler.reject_message(record.id, feedback)
                console.print("[yellow]Rejected.[/yellow]")
            else:
                console.print("[dim]Skipped.[/dim]")

        console.print(f"\n[bold green]Approved {approved} messages.[/bold green]")

        # Refresh Excel so Status and Shortlisted columns reflect review decisions
        path = self.export_to_excel()
        console.print(f"[bold cyan]Excel updated → {path}[/bold cyan]")
        return approved

    # ------------------------------------------------------------------ #
    # Phase 7: Send connection requests
    # ------------------------------------------------------------------ #

    def send_connections(self, limit: int = 20) -> int:
        approved = self._scheduler.list_by_status("approved")[:limit]
        if not approved:
            console.print("[yellow]No approved messages to send.[/yellow]")
            return 0

        sent = 0
        with LinkedInBrowser(headless=False) as browser:
            if not browser.is_logged_in():
                console.print("[yellow]Session expired — logging in with .env credentials...[/yellow]")
                if not browser.login(os.environ["LINKEDIN_EMAIL"], os.environ["LINKEDIN_PASSWORD"]):
                    console.print("[red]Login failed. Check LINKEDIN_EMAIL / LINKEDIN_PASSWORD in .env[/red]")
                    return 0
            for record in approved:
                console.print(f"\n  Sending to: [cyan]{record.profile_name}[/cyan]")
                result = browser.send_connection_request(record.profile_url, record.message)
                if result["success"]:
                    self._scheduler.mark_sent(record.id)
                    sent += 1
                    console.print("  [green]Sent.[/green]")
                else:
                    self._scheduler.mark_failed(record.id, result.get("error", "Unknown"))
                    console.print(f"  [red]Failed: {result.get('error')}[/red]")

        console.print(f"\n[bold green]Sent {sent} connection requests.[/bold green]")

        # Refresh Excel so Status column reflects sent/failed outcomes
        path = self.export_to_excel()
        console.print(f"[bold cyan]Excel updated → {path}[/bold cyan]")
        return sent
